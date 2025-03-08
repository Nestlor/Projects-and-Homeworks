import os
os.chdir('D:/dev/python/lesson11')
from openpyxl import Workbook, load_workbook

def create_file():
    data = [[1111], [2222], [3333]]
    for i, value in enumerate(data, 1):
        wb = Workbook()
        ws = wb.active
        ws.append(value)
        wb.save(f'file_{i}.xlsx')

def read_and_sort():
    all_data = []
    for i in range(1, 4):
        wb = load_workbook(f'file_{i}.xlsx')
        ws = wb.active
        for row in ws.iter_rows(values_only=True):
            all_data.extend(row)
    all_data.sort(reverse=True)
    return all_data

def write_sort_data(data):
    wb = Workbook()
    ws = wb.active

    for row_idx, value in enumerate(data, 1):
        ws.cell(row=row_idx, column=1, value=value)

    wb.save("sorted_data.xlsx")

create_file()
sorted_data = read_and_sort()
write_sort_data(sorted_data)